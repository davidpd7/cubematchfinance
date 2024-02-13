import os
import re
import logging

from PyQt6.QtWidgets import QFileDialog, QMessageBox, QLineEdit, QInputDialog
from PyQt6.QtCore import QStandardPaths

import PyPDF2
import pdfplumber
import pandas as pd
import subprocess
import openpyxl
import numpy as np
from datetime import datetime

class Model:

    def __init__(self):
       
        self.tab1 = self.Tab1(self)
        self.tab2 = self.Tab2(self)
        self.tab3 = self.Tab3(self)
        self.tab4 = self.Tab4(self)
        self.tab5 = self.Tab5(self)
        self.tab6 = self.Tab6(self)

    def sanitize_filename(self, filename):

        return re.sub(r'[<>:"/\\|?*]', '', filename)

    def rename(self, old_path, new_path):

        if os.path.exists(new_path):
            return None
        else:
            try:
                os.rename(src=old_path, dst=new_path)
            except Exception as e:
                None  
    
    class Tab1:

        def __init__(self, parent):

            self.parent = parent
        
        def browse(self, button):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def renaming_timesheets(self):

            try:
                file_names = self.fileName
                for pdf in file_names:
                    folder_path = os.path.dirname(pdf)
                    if pdf.endswith('.pdf'):
                        old_path = pdf
                        with pdfplumber.open(old_path) as pdf:
                            self.__renaming(pdf, folder_path, old_path)
            except Exception as e:
                    error_message = f"An error occurred while renaming info: {str(e)}"
                    QMessageBox.critical(None, "Error", error_message)

        def __renaming(self, pdf, folder_path, old_path):

            pages = pdf.pages
            pagina = pages[0]
            data = pagina.extract_text_lines()
            new_name = self.__set_name_timesheet(data)
            new_path = os.path.join(folder_path, new_name)
            pdf.close()
            self.parent.rename(new_path, old_path)
            
         
        def __set_name_timesheet(self, data):

            body = data[0]['text']
            name = data[1]['text'].split()[2:4]
            name = ' '.join(name)
            client = data[5]['text'].split()[1:]
            client = ' '.join(client)
            date = pd.to_datetime(data[2]['text'].replace('Timesheet period: ', '')[len('DD/MM/YYYY '):])
            new_name = f'{name} ({client}) {body} - {str(date.month_name())} {date.year}.pdf'
            new_name = self.parent.sanitize_filename(new_name)
            return new_name

    class Tab2:

        def __init__(self, parent):

            self.parent = parent
            self.fileName = []
            self.split_pdf_files = []
        
        def browse(self, button):
            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def split_pdf(self):

            try:
                files = self.fileName[0]
                self.split_pdf_files = []
                folder_path = os.path.dirname(files)
                output_dir = os.path.join(folder_path, 'Sales')
                os.makedirs(output_dir, exist_ok=True)
            
                with open(files, 'rb') as file:
                    pdf = PyPDF2.PdfReader(file)
                    total_pages = len(pdf.pages)
                    for page_number in range(total_pages):
                        pdf_writer = PyPDF2.PdfWriter()
                        pdf_writer.add_page(pdf.pages[page_number])
                        output_file_name = f'page_{page_number + 1}.pdf'
                        output_file_path = os.path.join(output_dir, output_file_name).replace("\\\\", "\\")
                        self.split_pdf_files.append(output_file_path)
                        with open(output_file_path, 'wb') as output_file:
                            pdf_writer.write(output_file)
            except Exception as e:
                error_message = f"An error occurred while splitting files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def renaming_invoices(self):
            try:
                pdf_files = self.split_pdf_files
                folder_path = os.path.dirname(pdf_files[0])
                for file in pdf_files:
                    if file.endswith('.pdf'):
                        old_path = os.path.join(folder_path, file)
                        with pdfplumber.open(old_path) as pdf:
                            page = pdf.pages[0].extract_tables()
                            new_name = self.__set_name_sales_invoices(page)
                            new_path = os.path.join(folder_path, new_name)
                            pdf.close()
                            self.parent.rename(old_path, new_path)
            except Exception as e:
                error_message = f"An error occurred while renaming files:\n{str(e)}"
                print(error_message)
                
        def __set_name_sales_invoices(self, page):

            try:
                invoice_number = page[0][2][1]
                name = page[1][2][0].split()[:2]
                name = ' '.join(name)
                date = pd.to_datetime(page[0][0][1][2:])
                new_name = f"{invoice_number} ({name}) - {str(date.month_name())} {date.year}.pdf"
                new_name = self.parent.sanitize_filename(new_name)
                return new_name
            except Exception as e:
                error_message = f"An error occurred while extracting info: {str(e)}"
                print(error_message)
                            
    class Tab3:

        def __init__(self, parent):

            self.parent = parent
        
        def browse(self, button):
            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
    
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
    
        def __docxtopdf(self, file):

            try:
                files_path = os.path.dirname(file)
                nombre_archivo, extension = os.path.splitext(file)
                file_path = os.path.join(files_path, file)
                pdf_path = os.path.join(files_path, f'{nombre_archivo}.pdf')
                with open(os.devnull, 'w') as nullfile:
                    subprocess.run(["docx2pdf", file_path, pdf_path], stdout=nullfile, stderr=nullfile, shell=True)
                os.remove(file_path)
                return pdf_path
            except Exception as e:
                error_message = f"An error occurred while converting to .pdf: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def browse_pb(self):

            try:
                self.pb, _ = QFileDialog.getOpenFileNames(None, 'Open File')
                
            except Exception as e:
                error_message = f"An error occurred  while getting Purchase Book: {str(e)}"
                print(error_message)
        
        def extract_purchasebook(self):

            try:
                self.df_pb = pd.read_excel(self.pb[0])
                return self.df_pb
            except Exception as e:
                error_message = f"No Purchase Book selected: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def renaming_contractors(self, files):

            pb = self.extract_purchasebook()
            files = self.__converter()
            for file in files:
                try:
                    old_path = os.path.dirname(file)
                    full_name = os.path.basename(file)
                    company = full_name.split('_')[0]
                    date = pd.to_datetime(full_name.split('_')[1])
                    company_trim, pb_trim = self.__trim_str(company, pb)
                    
                    if company_trim in pb_trim:
                        try:
                            new_name = self.__find_contractor_name(pb, company, date)
                            new_path = os.path.join(old_path, new_name)
                            os.rename(file, new_path)
                        except:
                            self.__renaming(file, old_path) 
                    else:
                        self.__renaming(file, old_path) 

                except Exception as e:
                    error_message = f"An error occurred while renaming: {str(e)}"
                    print(error_message)

        def __converter(self):
            
            try:
                new_file_names = []
                for file in self.fileName:
                    if file.endswith('.docx'):
                        file = self.__docxtopdf(file)
                    new_file_names.append(file)
                return new_file_names
            except:
                AttributeError      

        def __trim_str(self, company, pb):
         
            company_trim = company.lower().replace(" ", "")
            pb_trim = pb.iloc[:, 0].str.lower().str.replace(" ", "").values
            return company_trim, pb_trim

        def __find_contractor_name(self, pb, company, date):
         
            mask = pb.iloc[:, 0].str.lower().str.replace(" ", "") == company.lower().replace(" ", ""),pb.columns[1]
            contractor_name = pb.loc[mask].values[0]
            new_name = f"{company} ({contractor_name}) - {str(date.month_name())} {date.year}.pdf"
            new_name = self.parent.sanitize_filename(new_name)

            return new_name

        def __set_name_non_contractors(self, file_path, existing_files):
            full_name = os.path.basename(file_path)
            company_name = full_name.split('_')[0]
            date_str = full_name.split('_')[1]
            date = pd.to_datetime(date_str)
            formatted_date = f'{date.strftime("%B")} {date.year}'
            new_name = f'{company_name} - {formatted_date}.pdf'
            if new_name in existing_files:
                base_name, extension = os.path.splitext(new_name)
                new_name = f'{base_name} ({self.__get_next_unique_number(existing_files)}){extension}'
            
            return new_name

        def __get_next_unique_number(self, existing_files):

            secuential_number = 1
    
            while True:
                candidate = secuential_number
                if not any(f"_ {candidate}" in file for file in existing_files):
                    return candidate
                secuential_number += 1
        
        def __renaming(self, file, old_path):
            
            existing_files = os.listdir(old_path)
            new_name = self.__set_name_non_contractors(file, existing_files)
            new_path = os.path.join(old_path, new_name)
            os.rename(file, new_path)

        def renaming_noncontractors_other(self, files):
            
            files = self.__converter()
            if files is not None:
                old_path = os.path.dirname(files[0])
                for file in files:
                    try:
                        self.__renaming(file, old_path)
                    except Exception as e:
                        error_message = f"An error occurred while renaming: {str(e)}"
                        print(error_message)

    class Tab4:

        def __init__(self, parent):

            self.parent = parent
            
        def browse(self, button):
            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def salaries_extract(self):
            self.information = []
            try:
                file_names = self.fileName
                with pdfplumber.open(file_names[0]) as pdf:
                    page = pdf.pages[0].extract_tables()[0]
                    for fila in page:
                        if isinstance(fila, (list, dict)):
                            if isinstance(fila, dict):
                                if fila.get(0) is not None:
                                    self.information.append(fila)
                            elif isinstance(fila, list):
                                if fila[0].isdigit():
                                    self.information.append(fila)
            except Exception as e:
                error_message =f"An error occurred while extracting PDF info:{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
            return self.information

        def write_excel(self):
           
            try:
                information = self.salaries_extract()
                if information:
                    libro_excel, _ = QFileDialog.getOpenFileNames(None, "Select Files", r"C:\\")
                    archivo_excel = libro_excel[0]
                    libro_excel = openpyxl.load_workbook(archivo_excel)
                    replace_dict = {'.': '', ',': '.'}
                    
                    spreadsheet = libro_excel.active
            
                    for fila, datos in enumerate(information, start=12):
                        spreadsheet.cell(row=fila, column=2).value = datos[1]  
                        spreadsheet.cell(row=fila, column=3).value = datos[2] 
                        spreadsheet.cell(row=fila, column=4).value = datos[3].translate(str.maketrans(replace_dict))
                        spreadsheet.cell(row=fila, column=5).value = "CM Salary"
                    
                    libro_excel.save(archivo_excel)

            except Exception as e:
                error_message =f"An error occurred while writing on Excel: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)

    class Tab5:

        def __init__(self, parent):

            self.parent = parent

        def browse(self, button):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(button, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        

        def extract_clarity_details(self):
           
            try:
                df = self.clean_clarity()
                df_grouped = df.groupby(['Contractor PO Number', 'Surname', 'First Name']).agg(
                    Total_Hours=('Sum of Hours', 'sum'),
                    Shift=('Sum of Hours', pd.Series.mode ),
                    Total_Invoice=('Sum of Total', 'sum')).reset_index()
                df_grouped['Total Days'] = df_grouped['Total_Hours']/ df_grouped['Shift']
                self.df_result = df_grouped
            except Exception as e:
                error_message =f"Error while extracting Clarity information: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
            else:
                return  self.df_result

        
        def clean_clarity(self):
            try:
                file_names = self.fileName[0]
                self.folder_path = os.path.dirname(file_names)
                try:
                    df = pd.read_excel(file_names, engine='pyxlsb', skiprows=1, index_col=0)
                except:
                    df = pd.read_excel(file_names, skiprows=1, index_col=0)
                df.columns = df.iloc[df.index.get_loc('Resource ID')]
                mask = pd.notnull(df.columns)
                df = df.loc [:, mask] 
                df = df[~df.isna().any(axis=1)]
                df = df.iloc[1:]
                return df
            except Exception as e:
                error_message =f"Error while cleaning Clarity information: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def export_clarity_to_excel(self):
            try:
                self.df_result.to_excel(os.path.join(self.folder_path, "Clarity Details.xlsx"), index = False)
            except Exception as e:
                error_message =f"Error while exporting Clarity information: {str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
    
    class Tab6:

        def __init__(self, parent):

            self.parent = parent
            self.home_dir = QStandardPaths.standardLocations(QStandardPaths.StandardLocation.DesktopLocation)[0]
        
        def browse_database(self):

            try:
                self.fileName, _ = QFileDialog.getOpenFileNames(None, 'Open File')
            except Exception as e:
                error_message = f"An error occurred while browsing files:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        

        def __open_database(self):
            try:
                self.database = self.fileName[0]
                self.assignments = pd.read_excel(self.database, sheet_name='Assigments')
                self.sales = pd.read_excel(self.database , sheet_name='Sales')
                self.associates = pd.read_excel(self.database , sheet_name='Consultants')
            except Exception as e:
                error_message = f"An error occurred while opening the database:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)

        def __extract_sales_list(self):
            self.__open_database()
            try:
                active_assignments = self.assignments.loc[self.assignments['Status'] == "Active"]
                sales_list = active_assignments[['ID','Name', 'Client','Location', "Cost Rate"]].sort_values('Location')
                sales_list.rename(columns= {'ID':'Assignment ID'}, inplace=True)
                sales_list.rename(columns= {'Cost Rate':'Purchase Days'}, inplace=True)
                sales_list['Purchase Days'] = sales_list['Purchase Days'].apply(lambda x: "Permanent" if x == 0 else "")
                sales_list.reset_index(drop=True, inplace=True)
                sales_list = pd.concat([self.sales, sales_list], axis=0)
                ire_sales_list = sales_list.loc[sales_list['Location']== 'Ireland'].drop(columns='Location')
                bv_sales_list = sales_list.loc[sales_list['Location']== 'Netherlands'].drop(columns='Location')
                uk_sales_list = sales_list.loc[sales_list['Location']== 'United Kingdom'].drop(columns='Location')

                return ire_sales_list,  bv_sales_list, uk_sales_list
            except Exception as e:
                error_message = f"An error occurred while creating sales lists:\n{str(e)}"
                QMessageBox.critical(None, "Error", error_message)
        
        def export_sales_list(self):
            try:
                ire, bv, uk = self.__extract_sales_list()
                ire.to_excel(os.path.join(self.home_dir, 'CMIRE Sales List.xlsx'))
                bv.to_excel(os.path.join(self.home_dir, 'CMBV Sales List.xlsx'))
                uk.to_excel(os.path.join(self.home_dir, 'CMUK Sales List.xlsx')) 
            except:
                return

        
        def browse_sales_list(self):
            self.sales_list, _ = QFileDialog.getOpenFileNames(None, 'Open File')

        def reminder(self):
            self.__open_database()

            try:
                sales_list = self.sales_list[0]
                sheet_name = self.__get_sheet_name()
                sales_list = pd.read_excel(sales_list, sheet_name=sheet_name)
                reminder = pd.merge(sales_list, self.assignments, left_on='Assignment ID', right_on='ID', how='inner')
                reminder2 = pd.merge(reminder, self.associates, left_on='Associate.ID', right_on='ID', how='inner')
                Timesheet_reminder = reminder2.loc[reminder2['Days'].isna()][['Name', 'Email']].drop_duplicates()
                Purchases_reminder = reminder2.loc[reminder2['Purchase Days'].isna()][['Name', 'Email', 'Company Email']].drop_duplicates()
                Timesheet_reminder.to_csv(os.path.join(self.home_dir,'Timesheet_Reminder.csv'), index=False)
                Purchases_reminder.to_csv(os.path.join(self.home_dir,'Purchase_Reminder.csv'), index=False)
            except:
                return
            

        def __get_sheet_name(self):
            sheet_name_input, ok_pressed = QInputDialog.getText(None, "Sheet Name", "Enter the sheet name:")

            if ok_pressed:
                return sheet_name_input

            return ""

        def clarity_reminder(self):
            self.__open_database()
            try:
                Clarity = self.assignments.loc[(self.assignments['Status'] == "Active") & (self.assignments['Client'] == 'Bank of Ireland (IRE)')]
                ActiveBOI = pd.merge(Clarity, self.associates, left_on ='Associate.ID', right_on='ID', how='inner')
                ActiveBOI = ActiveBOI.loc[ActiveBOI['Comments'] != 'No clarity'][['Name_x', 'Email']].drop_duplicates()
                ActiveBOI.to_csv(os.path.join(self.home_dir,'Clarity Reminder.csv'))
            except Exception as e:
                    print(f"Error reading Excel file: {e}")
        
        def associates_view(self, source:str ):
            try:
                self.__open_database()
                return self.associates
            except:
                pass
        
        def assignments_view(self, source: str):
            try:
                self.__open_database()
                return self.assignments
            except:
                pass
                    
        

